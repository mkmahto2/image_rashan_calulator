import cv2
import pytesseract
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

# Set the path for Tesseract executable if necessary
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Uncomment and modify this line if needed

# Function to preprocess the image for better OCR results
def preprocess_image(image_path):
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    
    # Resize the image
    img = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    
    # Apply Gaussian Blur
    img = cv2.GaussianBlur(img, (5, 5), 0)
    
    # Apply Thresholding
    _, img = cv2.threshold(img, 150, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    
    # Save the preprocessed image for debugging
    preprocessed_path = "preprocessed_image.png"
    cv2.imwrite(preprocessed_path, img)
    print(f"Preprocessed image saved as {preprocessed_path}")

    return preprocessed_path  # Return the path instead of the image array

# Function to extract text from the preprocessed image
def image_to_text(image_path):
    img = preprocess_image(image_path)
    custom_config = r'--oem 3 --psm 6'  # PSM 6 for uniform blocks of text
    extracted_text = pytesseract.image_to_string(cv2.imread(img), config=custom_config)
    
    print(f"Raw Extracted Text: \n{extracted_text}")  # Print raw text for debugging
    
    if not extracted_text.strip():
        print("No text extracted from the image.")
    return extracted_text

# Function to clean and structure the text into a list of tuples (Product Name, Quantity)
def clean_and_structure_text(text):
    product_list = text.splitlines()
    structured_list = []
    
    for line in product_list:
        line = line.strip()
        if line:
            parts = line.split()  # Assuming space-separated values
            if len(parts) >= 1:  # At least one part should be present (the quantity)
                quantity = parts[-1].strip()  # Last part as the quantity
                product_name = ' '.join(parts[:-1]).strip()  # Join all parts except the last one as product name
                structured_list.append((product_name, quantity))
            else:
                print(f"Line does not contain enough parts: {line}")  # Debugging output
    return structured_list

# Function to save the structured list of products and the processed image in the same Excel file
def save_to_excel(structured_list, processed_image_path, excel_path):
    wb = Workbook()
    
    # Save product list in the first sheet
    ws1 = wb.active
    ws1.title = "Product List"
    
    # # Write headers
    ws1['A1'] = 'Product Name'
    ws1['B1'] = 'Quantity'
    
    # Write product data to the sheet
    for idx, (product_name, quantity) in enumerate(structured_list, start=2):
        ws1.cell(row=idx, column=1).value = product_name
        ws1.cell(row=idx, column=2).value = quantity
    
    # Create a new sheet for the processed image
    ws2 = wb.create_sheet(title="Processed Image")
    
    img = ExcelImage(processed_image_path)
    img.anchor = 'A1'  # Adjust the position as needed
    ws2.add_image(img)
    
    # Save the workbook
    wb.save(excel_path)
    print(f"Data and processed image saved to {excel_path}")

# Main function to process the image and save it to an Excel file
def process_image_to_excel(image_path, excel_path):
    text_output = image_to_text(image_path)
    print(f"Extracted Text: \n{text_output}")  # Print the extracted text for verification
    
    structured_list = clean_and_structure_text(text_output)
    
    # Print structured data
    print("\nProduct Name and Quantity:")
    for product_name, quantity in structured_list:
        print(f"Product: {product_name}, Quantity: {quantity}")  # Print each product name and quantity
    
    # Save structured data and processed image to the same Excel file
    processed_image_path = preprocess_image(image_path)  # Call this again to ensure the latest image is processed
    save_to_excel(structured_list, processed_image_path, excel_path)

# Example usage
image_path = 'product.png'  # Replace with your actual image path
excel_path = 'product_list_with_image.xlsx'

process_image_to_excel(image_path, excel_path)



import pandas as pd

def clean_product_names(excel_path, cleaned_excel_path, characters_to_remove):
    # Read the product list Excel file
    product_df = pd.read_excel(excel_path, sheet_name='Product List')
    
    # Print original data for debugging
    print("Original Product Data:\n", product_df.head())  
    
