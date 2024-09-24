import cv2
import pytesseract
from PIL import Image





# Set the path to tesseract.exe (replace with the correct path)
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'





import cv2
from PIL import Image
import pytesseract
from openpyxl import Workbook

# Set the path for Tesseract executable if necessary
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Uncomment and modify this line if needed

# Function to preprocess the image for better OCR results
def preprocess_image(image_path):
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    
    # Resize the image
    img = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    
    # Apply Gaussian Blur
    img = cv2.GaussianBlur(img, (5, 5), 0)
    
    # Apply Thresholding
    _, img = cv2.threshold(img, 150, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    
    return img

# Function to extract text from the preprocessed image
def image_to_text(image_path):
    img = preprocess_image(image_path)
    
    # Save the preprocessed image temporarily for debugging (optional)
    # cv2.imwrite("preprocessed_image.png", img)

    custom_config = r'--oem 3 --psm 6'  # PSM 6 for uniform blocks of text
    extracted_text = pytesseract.image_to_string(img, config=custom_config)
    return extracted_text

# Function to clean and structure the text into a list of tuples (Product Name, Rate, Quantity)
def clean_and_structure_text(text):
    product_list = text.splitlines()
    structured_list = []
    
    for line in product_list:
        line = line.strip()
        if line:
            # Split the line by spaces or a defined separator. Adjust based on the actual format.
            parts = line.split()  # Assuming space-separated values; change as necessary.
            if len(parts) >= 3:
                product_name = " ".join(parts[:-2]).strip()  # Join all but the last two parts for the product name
                rate = parts[-2].strip()  # Second to last part as the rate
                quantity = parts[-1].strip()  # Last part as the quantity
                structured_list.append((product_name, rate, quantity))
            else:
                print(f"Line does not contain enough parts: {line}")  # Debugging output
    return structured_list

# Function to save the structured list of products into an Excel file
def save_list_to_excel(structured_list, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Product List"
    
    # # Write headers
    # ws['A1'] = 'Product Name'
    # ws['B1'] = 'Rate'
    # ws['C1'] = 'Quantity'
    
    # Write product data to the sheet
    for idx, (product_name, rate, quantity) in enumerate(structured_list, start=2):
        ws.cell(row=idx, column=1).value = product_name
        ws.cell(row=idx, column=2).value = rate
        ws.cell(row=idx, column=3).value = quantity
    
    # Save the workbook
    wb.save(excel_path)

# Main function to process the image and save it to an Excel file
def process_image_to_excel(image_path, excel_path):
    text_output = image_to_text(image_path)
    print(f"Extracted Text: \n{text_output}")  # Print the extracted text for verification
    
    structured_list = clean_and_structure_text(text_output)
    
    # Debugging: Check if structured data is populated
    if not structured_list:
        print("No structured data to save.")
        return
    
    # Save structured data to Excel
    save_list_to_excel(structured_list, excel_path)
    print(f"Data saved to {excel_path}")

# Example usage
image_path = 'product.png'  # Replace with your actual image path
excel_path = 'product_list_output.xlsx'

process_image_to_excel(image_path, excel_path)
