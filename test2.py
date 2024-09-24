from PIL import Image
import pytesseract
from openpyxl import Workbook



# Set the path to tesseract.exe (replace with the correct path)
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

from PIL import Image
import pytesseract
from openpyxl import Workbook

# Function to extract text from an image
def image_to_text(image_path):
    # Open the image using PIL
    img = Image.open(image_path)
    
    # Use Tesseract to extract text
    extracted_text = pytesseract.image_to_string(img)
    
    return extracted_text

# Function to clean and structure the text into a list of tuples (Product Name, Rate, Quantity)
def clean_and_structure_text(text):
    # Split the text into lines
    product_list = text.splitlines()
    
    # Clean up and split each line into product name, rate, and quantity
    structured_list = []
    for line in product_list:
        line = line.strip()
        if line:
            # Assuming the format is "Product Name - Rate - Quantity"
            parts = line.split('-')  # Split by hyphen (change the separator based on your image format)
            if len(parts) == 3:
                product_name = parts[0].strip()
                rate = parts[1].strip()
                quantity = parts[2].strip()
                structured_list.append((product_name, rate, quantity))
    
    return structured_list

# Function to save the structured list of products into an Excel file with columns for Product Name, Rate, and Quantity
def save_list_to_excel(structured_list, excel_path):
    # Create a new Excel workbook
    wb = Workbook()
    
    # Select the active worksheet
    ws = wb.active
    ws.title = "Product List"  # Optionally name the sheet

    # Write the headers
    ws['A1'] = 'Product Name'
    ws['B1'] = 'Rate'
    ws['C1'] = 'Quantity'

    # Save each product's data in new rows (starting from row 2, as row 1 is for headers)
    for idx, (product_name, rate, quantity) in enumerate(structured_list, start=2):
        ws.cell(row=idx, column=1).value = product_name   # Product Name in Column A
        ws.cell(row=idx, column=2).value = rate           # Rate in Column B
        ws.cell(row=idx, column=3).value = quantity       # Quantity in Column C
    
    # Save the workbook
    wb.save(excel_path)

# Main function to process the image and save it to an Excel file
def process_image_to_excel(image_path, excel_path):
    # Extract text from the image
    text_output = image_to_text(image_path)

    # Clean and structure the extracted text into a list of (Product Name, Rate, Quantity) tuples
    structured_list = clean_and_structure_text(text_output)

    # Save the structured list to Excel
    save_list_to_excel(structured_list, excel_path)

    print(f"Product list with rates and quantities has been saved to {excel_path}")

# Example usage
image_path = 'product.png'  # Replace with the actual image file path
excel_path = 'product_list_output.xlsx'

# Process the image and generate the Excel file
process_image_to_excel(image_path, excel_path)
