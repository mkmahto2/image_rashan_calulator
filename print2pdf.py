from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def read_excel(file_path):
    # Load the workbook and select the active worksheet
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Read product data from the Excel file
    products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        product_name = row[0]
        rate = row[1]
        quantity = row[2]
        products.append((product_name, rate, quantity))
    
    return products

def create_pdf_bill(file_path, products):
    # Create a canvas object
    c = canvas.Canvas(file_path, pagesize=letter)
    
    # Set the title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(100, 750, "Final Bill")
    
    # Set the headers
    c.setFont("Helvetica-Bold", 12)
    c.drawString(100, 720, "Product Name")
    c.drawString(300, 720, "Rate")
    c.drawString(400, 720, "Quantity")
    
    # Set the font for the product details
    c.setFont("Helvetica", 12)
    
    # Starting position for the product details
    y_position = 700

    # Loop through products and add them to the PDF
    for product_name, rate, quantity in products:
        c.drawString(100, y_position, str(product_name))
        c.drawString(300, y_position, str(rate))
        c.drawString(400, y_position, str(quantity))
        y_position -= 20  # Move down for the next item
    
    # Save the PDF
    c.save()

# Main function to convert Excel to PDF
def convert_excel_to_pdf(excel_path, pdf_path):
    products = read_excel(excel_path)
    create_pdf_bill(pdf_path, products)
    print(f"PDF bill has been created at {pdf_path}")

# Example usage
excel_path = 'final_bill.xlsx'  # Replace with your Excel file path
pdf_path = 'final_bill_from_excel.pdf'

# Convert the Excel file to PDF
convert_excel_to_pdf(excel_path, pdf_path)
